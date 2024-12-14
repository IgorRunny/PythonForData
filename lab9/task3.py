from openpyxl import load_workbook
from openpyxl.chart import Reference, PieChart

file_path = 'table.xlsx'
workbook = load_workbook(file_path)
sheet = workbook.active

department_salaries = {}
for row_index in [4, 9, 11]:
    department_name = sheet[row_index][2].value.split(' ')[0]
    department_salaries[department_name] = sheet[row_index][5].value

summary_sheet = workbook.create_sheet(title="Summary")
summary_sheet.append(["Отдел", "Сумма зарплаты"])
for department, salary in department_salaries.items():
    summary_sheet.append([department, salary])

chart = PieChart()
chart.title = "Распределение зарплат по отделам"
labels = Reference(summary_sheet, min_col=1, min_row=2, max_row=len(department_salaries) + 1)
data = Reference(summary_sheet, min_col=2, min_row=1, max_row=len(department_salaries) + 1)
chart.add_data(data, titles_from_data=True)
chart.set_categories(labels)
sheet.add_chart(chart, "M1")

output_file_path = 'table.xlsx'
workbook.save(output_file_path)